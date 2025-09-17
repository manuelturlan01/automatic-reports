#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tickets_parser.py
-----------------
Extrae campos de PDFs de tickets (formato Service Desk similar al que enviaste) y
genera un Excel en español con columnas útiles + tiempos calculados.

Uso:
  python tickets_parser.py --pdf_dir "C:/ruta/a/pdfs" --out "C:/ruta/salida/Tickets.xlsx"

Opcionales:
  --pattern "Ticket-*.pdf"   (por defecto procesa TODOS los .pdf del directorio: "*.pdf")
  --timezone "America/Asuncion"
  --dedupe                   (desduplicar por N° Ticket quedando con la fila más completa)
  --dump_text                (guardar debug_texts con el texto leído)

Requisitos:
  pip install -U pandas PyMuPDF PyPDF2 openpyxl
"""
import os, re, sys, argparse, time, unicodedata
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from zoneinfo import ZoneInfo
from glob import glob

try:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False
    get_column_letter = None  # type: ignore[assignment]
    DataValidation = None  # type: ignore[assignment]
    DefinedName = None  # type: ignore[assignment]

CBSA_DEPARTMENTS: List[str] = [
    "Mesa de Dinero",
    "IT",
    "Comercial",
    "Administración",
    "Desarrollo de Negocios",
    "Finanzas Corporativas",
    "Riesgos",
    "Tesorería",
    "Operaciones",
    "Contabilidad",
    "Cumplimiento",
]

AREA_CBSA = "CBSA"
AREA_FONDOS = "Fondos"
AREA_OPTIONS = [AREA_CBSA, AREA_FONDOS]

PRIORITY_OPTIONS: List[str] = [
    "Baja",
    "Media",
    "Alta",
    "Crítica",
]

# ============== Backends de extracción de texto ==============
try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except Exception:
    HAS_PYMUPDF = False

try:
    from PyPDF2 import PdfReader
    HAS_PYPDF2 = True
except Exception:
    HAS_PYPDF2 = False

def log(msg: str):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)

def extract_text_pymupdf(pdf_path: str) -> str:
    doc = fitz.open(pdf_path)
    chunks = []
    for i, page in enumerate(doc, start=1):
        try:
            chunks.append(page.get_text("text") or "")
        except Exception as e:
            log(f"   - PyMuPDF page {i} error: {e}")
    return "\n".join(chunks).strip()

def extract_text_pypdf2(pdf_path: str) -> str:
    reader = PdfReader(pdf_path)
    chunks = []
    for i, page in enumerate(reader.pages, start=1):
        try:
            txt = page.extract_text() or ""
            chunks.append(txt)
        except Exception as e:
            log(f"   - PyPDF2 page {i} error: {e}")
    return "\n".join(chunks).strip()

def extract_text(pdf_path: str) -> str:
    if HAS_PYMUPDF:
        txt = extract_text_pymupdf(pdf_path)
        if len(txt) > 80:
            return txt
    if HAS_PYPDF2:
        txt = extract_text_pypdf2(pdf_path)
        if len(txt) > 80:
            return txt
    return ""

# ============== Limpieza y patrones ==============
FOOTER_RE = re.compile(r"Ticket\s*#\d+\s+printed by.*?Page\s*\d+", re.I | re.S)
WEEKDAY_LINE = re.compile(r"^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\b.*$", re.I | re.M)

STAMP_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4}),?\s+(\d{1,2}):(\d{2})(?::(\d{2}))?\s*([AP]M)?", re.I)
TS_LINE_RE = STAMP_RE

LABELY = re.compile(r"^[A-Za-zÁÉÍÓÚÜÑáéíóúüñ/ .]{2,60}\s*:\s*")

HEADER_FIELD_VARIANTS: Dict[str, List[str]] = {
    "Status": ["Status", "Estado"],
    "Name": ["Name", "Nombre"],
    "Priority": ["Priority", "Prioridad"],
    "Department": ["Department", "Departamento"],
    "Create Date": ["Create Date", "Fecha de creación", "Fecha de creacion"],
}

ADDITIONAL_HEADER_LABELS = [
    "Email",
    "Phone",
    "Source",
    "Ticket Details",
    "Urgency",
    "Related client number or user",
    "Related client",
    "Cliente relacionado",
    "Cliente/usuario relacionado",
]


def normalize_label_text(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace(":", " ").replace("#", " ")
    return " ".join(s.lower().split())


def label_to_pattern(label: str) -> str:
    parts = [re.escape(part) for part in label.strip().split() if part]
    if not parts:
        return ""
    return r"\s+".join(parts)


HEADER_FIELD_INLINE_PATTERNS: Dict[str, List[re.Pattern[str]]] = {
    key: [
        re.compile(rf"^(?:{label_to_pattern(label)})\s*[:\-]?\s*(.+)$", re.I)
        for label in variants
    ]
    for key, variants in HEADER_FIELD_VARIANTS.items()
}

HEADER_FIELD_NORMALIZED: Dict[str, List[str]] = {
    key: [normalize_label_text(label) for label in variants]
    for key, variants in HEADER_FIELD_VARIANTS.items()
}

ALL_KNOWN_HEADER_LABELS = set()
for variants in HEADER_FIELD_NORMALIZED.values():
    ALL_KNOWN_HEADER_LABELS.update(variants)
ALL_KNOWN_HEADER_LABELS.update(normalize_label_text(lbl) for lbl in ADDITIONAL_HEADER_LABELS)

HEADER_SEARCH_LIMIT = 120

def clean_text(text: str) -> str:
    text = FOOTER_RE.sub("", text)
    text = WEEKDAY_LINE.sub("", text)
    return text

def find_line_idx(lines: List[str], predicate, start=0, limit=None) -> Optional[int]:
    end = len(lines) if limit is None else min(len(lines), start+limit)
    for i in range(start, end):
        if predicate(lines[i]):
            return i
    return None

def is_related_label(s: str) -> bool:
    t = " ".join(s.strip().lower().split())
    return t.endswith(":") and (
        "related client number or user" in t or "related client" in t or
        "cliente relacionado" in t or "usuario relacionado" in t or
        "cliente/usuario relacionado" in t
    )

def extract_title_after_urgency(cleaned: str) -> str:
    lines = cleaned.splitlines()
    td = find_line_idx(lines, lambda ln: ln.strip().lower().startswith("ticket details"), 0, 300)
    idx = (td + 1) if td is not None else 0

    skip_next = False
    while idx < len(lines):
        raw = lines[idx]
        idx += 1
        s = raw.strip()
        if not s:
            continue

        if skip_next:
            skip_next = False
            continue

        # Saltar bloques de "Urgency" que aparecen antes del título
        if re.match(r"^\s*Urgency", s, re.I):
            skip_next = True  # su valor está en la siguiente línea
            continue

        # Saltar etiquetas de cliente relacionado y su valor
        if is_related_label(raw):
            skip_next = True
            continue

        # Ignorar líneas que parezcan etiquetas del encabezado u otros campos
        if LABELY.match(s):
            continue

        # Ignorar líneas que en realidad son un timestamp de la conversación
        if TS_LINE_RE.search(s):
            continue

        return s

    return ""

def is_auto(author: str, body: str) -> bool:
    blob = (author + " " + body).lower()
    auto_keys = ["mail delivery subsystem","delivery status notification","mailer-daemon","postmaster","do not reply","no-reply"]
    return any(k in blob for k in auto_keys)

def extract_thread_entries(full_text: str) -> List[Dict[str,str]]:
    entries = []
    matches = list(TS_LINE_RE.finditer(full_text))
    for i, m in enumerate(matches):
        line_end   = full_text.find("\n", m.end())
        if line_end == -1:
            line_end = len(full_text)
        header_line = full_text[m.start():line_end]
        author = header_line[m.end()-m.start():].strip()
        body = full_text[line_end:matches[i+1].start()] if i+1 < len(matches) else full_text[line_end:]
        entries.append({"stamp": m.group(0), "author": author, "body": body.strip()})
    return entries

def fallback_last_author_from_tail(cleaned: str) -> str:
    lines = [ln.strip() for ln in cleaned.splitlines()]
    for s in reversed(lines):
        if not s:
            continue
        if TS_LINE_RE.search(s):
            continue
        if LABELY.match(s):
            continue
        if s.lower() in ("ticket details","urgency:"):
            continue
        return s
    return ""

def parse_timestamp_ddmm(s: str, tz: ZoneInfo) -> Optional[datetime]:
    m = STAMP_RE.search(s or "")
    if not m:
        return None
    d1, d2, yyyy, hh, mm, ss, ampm = m.groups()
    d, mo = int(d1), int(d2)
    year = int(yyyy) + (2000 if int(yyyy) < 100 else 0)
    hour, minute = int(hh), int(mm)
    sec = int(ss) if ss else 0
    if ampm:
        if ampm.upper() == "PM" and hour != 12: hour += 12
        if ampm.upper() == "AM" and hour == 12: hour = 0
    try:
        return datetime(year, mo, d, hour, minute, sec, tzinfo=tz)
    except Exception:
        return None

def parse_timestamp_mmdd(s: str, tz: ZoneInfo) -> Optional[datetime]:
    m = STAMP_RE.search(s or "")
    if not m:
        return None
    d1, d2, yyyy, hh, mm, ss, ampm = m.groups()
    mo, d = int(d1), int(d2)
    year = int(yyyy) + (2000 if int(yyyy) < 100 else 0)
    hour, minute = int(hh), int(mm)
    sec = int(ss) if ss else 0
    if ampm:
        if ampm.upper() == "PM" and hour != 12: hour += 12
        if ampm.upper() == "AM" and hour == 12: hour = 0
    try:
        return datetime(year, mo, d, hour, minute, sec, tzinfo=tz)
    except Exception:
        return None

MAX_FUTURE_DELTA = timedelta(days=365)


def parse_pdf_timestamp(s: str, now: datetime, tz: ZoneInfo) -> Optional[datetime]:
    candidates = []
    for parse_fn in (parse_timestamp_ddmm, parse_timestamp_mmdd):
        dt = parse_fn(s, tz)
        if not dt:
            continue
        if dt > now + MAX_FUTURE_DELTA:
            continue
        candidates.append(dt)

    if not candidates:
        return None

    non_future = [dt for dt in candidates if dt <= now]
    pool = non_future if non_future else candidates
    return min(pool, key=lambda dt: abs((dt - now).total_seconds()))

def is_open_status(status: str) -> bool:
    if not isinstance(status, str):
        return True
    st = status.strip().lower()
    closed_words = ["closed","resolved","solved","cerrado","resuelto","completado","finalizado","finalizada"]
    return not any(w in st for w in closed_words)

def parse_header_block(cleaned: str, lines: List[str]) -> Dict[str,str]:
    header = {
        "Ticket Number": "",
        "Status": "",
        "Name": "",
        "Priority": "",
        "Department": "",
        "Create Date": "",
    }

    ticket_match = re.search(r"Ticket\s*#(\d+)", cleaned)
    if ticket_match:
        header["Ticket Number"] = ticket_match.group(1).strip()

    stripped_lines = [ln.strip() for ln in lines]
    normalized_lines = [normalize_label_text(ln) for ln in stripped_lines]

    def find_value(field: str) -> str:
        inline_patterns = HEADER_FIELD_INLINE_PATTERNS.get(field, [])
        normalized_variants = HEADER_FIELD_NORMALIZED.get(field, [])

        limit = min(len(stripped_lines), HEADER_SEARCH_LIMIT)
        for idx in range(limit):
            raw_line = stripped_lines[idx]
            if not raw_line:
                continue

            for pat in inline_patterns:
                m = pat.match(raw_line)
                if m:
                    value = m.group(1).strip()
                    if value and normalize_label_text(value) not in ALL_KNOWN_HEADER_LABELS:
                        return value

            if normalized_lines[idx] in normalized_variants:
                for j in range(idx + 1, min(len(stripped_lines), idx + 6)):
                    candidate = stripped_lines[j].strip()
                    if not candidate:
                        continue
                    if normalize_label_text(candidate) in ALL_KNOWN_HEADER_LABELS:
                        break
                    return candidate
        return ""

    for field in ("Status", "Name", "Priority", "Department", "Create Date"):
        header[field] = find_value(field)

    return header

EMPTY_PARSE_RESULT: Dict[str, str] = {
    "N° Ticket": "",
    "Título del ticket": "",
    "Estado BW": "",
    "Prioridad": "",
    "Departamento": "",
    "Fecha de creación": "",
    "Autor": "",
    "Última respuesta por": "",
    "Última respuesta el": "",
}

def parse_pdf(pdf_path: str, tz: ZoneInfo, now: datetime) -> Dict[str,str]:
    raw = extract_text(pdf_path)
    if not raw or len(raw) < 80:
        return EMPTY_PARSE_RESULT.copy()
    cleaned = clean_text(raw)
    lines = cleaned.splitlines()
    hdr = parse_header_block(cleaned, lines)
    title = extract_title_after_urgency(cleaned)

    entries = extract_thread_entries(cleaned)
    last_by, last_at = "", ""
    first_author = ""
    for e in entries:
        if e["author"] and not is_auto(e["author"], e["body"]):
            if not first_author:
                first_author = e["author"]
            last_by, last_at = e["author"], e["stamp"]

    if not last_by:
        last_by = fallback_last_author_from_tail(cleaned)

    if not first_author:
        first_author = fallback_last_author_from_tail(cleaned)

    result = EMPTY_PARSE_RESULT.copy()
    result.update({
        "N° Ticket": hdr.get("Ticket Number", "").strip(),
        "Título del ticket": title.strip(),
        "Estado BW": hdr.get("Status", "").strip(),
        "Prioridad": hdr.get("Priority", "").strip(),
        "Departamento": hdr.get("Department", "").strip(),
        "Fecha de creación": hdr.get("Create Date", "").strip(),
        "Autor": first_author.strip(),
        "Última respuesta por": last_by.strip(),
        "Última respuesta el": last_at.strip(),
    })
    return result


def normalize_output_path(raw_path: str) -> str:
    """Return a sanitized Excel output path."""

    path = (raw_path or "").strip()
    if not path:
        raise ValueError("La ruta de salida no puede estar vacía.")

    allowed_exts = (".xlsx", ".xlsm")
    base, ext = os.path.splitext(path)
    ext_lower = ext.lower()

    if ext_lower in allowed_exts:
        return path

    for allowed_ext in allowed_exts:
        if ext_lower.startswith(allowed_ext):
            corrected = base + allowed_ext
            if corrected != path:
                log(
                    "Aviso: la ruta de salida se ajustó a "
                    f"'{corrected}' (valor original: '{path}')."
                )
            return corrected

    if not ext:
        corrected = path + ".xlsx"
        log(
            "Aviso: la ruta de salida no incluía extensión; "
            f"se utilizará '{corrected}'."
        )
        return corrected

    corrected = base + ".xlsx"
    log(
        "Aviso: la extensión de salida era inválida; "
        f"se utilizará '{corrected}'."
    )
    return corrected


def main():
    ap = argparse.ArgumentParser(description="PDF tickets -> Excel (ES)")
    ap.add_argument("--pdf_dir", required=True, help="Carpeta con PDFs")
    ap.add_argument("--out", required=True, help="Ruta de salida del Excel")
    ap.add_argument("--pattern", default="*.pdf", help="Patrón de archivos (glob)")
    ap.add_argument("--timezone", default="America/Asuncion", help="Zona horaria")
    args = ap.parse_args()

    tz = ZoneInfo(args.timezone)
    now = datetime.now(tz)

    files = sorted(glob(os.path.join(args.pdf_dir, args.pattern)))
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files:
        log("No se encontraron PDFs.")
        sys.exit(1)

    log(f"Procesando {len(files)} PDF(s)…")
    rows = []
    for path in files:
        row = parse_pdf(path, tz, now)
        rows.append(row)

    df = pd.DataFrame(rows)

    manual_columns = ["Prioridad", "Área", "Departamento"]

    df["Prioridad"] = ""

    rename_map = {}
    if "N° Ticket" in df.columns:
        rename_map["N° Ticket"] = "N Ticket"
    if "Estado" in df.columns:
        rename_map["Estado"] = "Estado BW"
    if rename_map:
        df = df.rename(columns=rename_map)

    estado_bw_series = None
    if "Estado BW" in df.columns:
        estado_bw_series = df["Estado BW"].copy()
        df = df.drop(columns=["Estado BW"])

    if "Área" not in df.columns:
        df["Área"] = ""
    else:
        df["Área"] = df["Área"].fillna("").astype(str).str.strip()

    if "Departamento" not in df.columns:
        df["Departamento"] = ""
    else:
        df["Departamento"] = df["Departamento"].fillna("").astype(str).str.strip()

    def normalize_department(row):
        area = (row.get("Área") or "").strip()
        dept = (row.get("Departamento") or "").strip()
        if not dept:
            return ""
        if not area:
            return ""
        if area == AREA_FONDOS:
            return "-" if dept == "-" else ""
        if area == AREA_CBSA:
            return dept if dept in CBSA_DEPARTMENTS else ""
        return dept

    df["Departamento"] = df.apply(normalize_department, axis=1)

    desired_order = [
        "N Ticket",
        "Título del ticket",
        "Autor",
        "Prioridad",
        "Área",
        "Departamento",
    ]
    leading_cols = [col for col in desired_order if col in df.columns]
    remaining_cols = [col for col in df.columns if col not in leading_cols]
    df = df.loc[:, leading_cols + remaining_cols]

    if "N Ticket" in df.columns:
        df["N Ticket"] = df["N Ticket"].fillna("").astype(str)
        df = df.drop_duplicates(subset=["N Ticket"], keep="last").reset_index(drop=True)

    df = df.fillna("")

    sheet_name = "Tickets"

    existing_df = None
    out_path = normalize_output_path(args.out)
    if os.path.exists(out_path):
        try:
            existing_df = pd.read_excel(out_path, sheet_name=sheet_name).fillna("")
        except Exception as exc:
            log(f"Aviso: no se pudo leer el Excel existente '{out_path}': {exc}")
            existing_df = None

    if existing_df is not None:
        if "N Ticket" in existing_df.columns:
            existing_df["N Ticket"] = existing_df["N Ticket"].fillna("").astype(str)
            existing_df = existing_df.drop_duplicates(subset=["N Ticket"], keep="first").reset_index(drop=True)

        for col in df.columns:
            if col not in existing_df.columns:
                existing_df[col] = ""

        existing_df = existing_df.reindex(columns=df.columns, fill_value="")
        df = df.reindex(columns=existing_df.columns, fill_value="")

        idx_by_ticket = {
            ticket: idx for idx, ticket in existing_df["N Ticket"].items() if ticket
        }

        rows_to_append = []
        for _, row in df.iterrows():
            ticket = row.get("N Ticket", "")
            if ticket and ticket in idx_by_ticket:
                row_idx = idx_by_ticket[ticket]
                for col in df.columns:
                    if col in manual_columns:
                        continue
                    existing_df.at[row_idx, col] = row[col]
            else:
                rows_to_append.append(row)

        if rows_to_append:
            new_rows_df = pd.DataFrame(rows_to_append, columns=df.columns)
            existing_df = pd.concat([existing_df, new_rows_df], ignore_index=True)

        df = existing_df

    df = df.fillna("")

    def to_local_naive_timestamp(value) -> pd.Timestamp:
        dt: Optional[datetime] = None
        if isinstance(value, pd.Timestamp):
            dt = value.to_pydatetime()
        elif isinstance(value, datetime):
            dt = value
        elif isinstance(value, str) and value.strip():
            dt = parse_pdf_timestamp(value, now, tz)
        if dt is None:
            return pd.NaT  # type: ignore[return-value]
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=tz)
        else:
            dt = dt.astimezone(tz)
        return pd.Timestamp(dt.replace(tzinfo=None))

    datetime_columns = [
        col for col in ("Fecha de creación", "Última respuesta el") if col in df.columns
    ]
    for col in datetime_columns:
        df[col] = df[col].apply(to_local_naive_timestamp)

    now_local = now.replace(tzinfo=None)
    now_timestamp = pd.Timestamp(now_local)
    zero_delta = pd.Timedelta(0)

    if "Última respuesta el" in df.columns:
        last_response_series = df["Última respuesta el"]
        wait_deltas = now_timestamp - last_response_series
        wait_deltas = wait_deltas.where(last_response_series.notna(), pd.NaT)
        wait_deltas = wait_deltas.where(wait_deltas > zero_delta, pd.NaT)
        df["Tiempo parado desde la última respuesta"] = wait_deltas

    if "Fecha de creación" in df.columns:
        creation_series = df["Fecha de creación"]
        open_status_mask = pd.Series(True, index=df.index)
        if estado_bw_series is not None:
            open_status_mask = estado_bw_series.apply(is_open_status)
            open_status_mask = open_status_mask.reindex(df.index, fill_value=True)
        open_deltas = now_timestamp - creation_series
        open_deltas = open_deltas.where(creation_series.notna(), pd.NaT)
        open_deltas = open_deltas.where(open_status_mask, pd.NaT)
        open_deltas = open_deltas.where(open_deltas > zero_delta, pd.NaT)
        df["Tiempo abierto (si sigue abierto)"] = open_deltas

    duration_columns = [
        col
        for col in (
            "Tiempo parado desde la última respuesta",
            "Tiempo abierto (si sigue abierto)",
        )
        if col in df.columns
    ]

    if duration_columns:
        seconds_per_day = 24 * 60 * 60
        for col in duration_columns:
            timedelta_series = pd.to_timedelta(df[col], errors="coerce")
            df[col] = timedelta_series.dt.total_seconds() / seconds_per_day

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    if not HAS_OPENPYXL:
        raise RuntimeError(
            "Se requiere openpyxl para generar el Excel con listas desplegables. "
            "Instalá openpyxl (pip install openpyxl) e intentá nuevamente."
        )

    data_row_start = 2
    data_row_end = max(len(df) + 1, 200)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        wb = writer.book
        ws = writer.sheets[sheet_name]

        priority_idx = df.columns.get_loc("Prioridad") + 1
        area_idx = df.columns.get_loc("Área") + 1
        dept_idx = df.columns.get_loc("Departamento") + 1
        priority_col = get_column_letter(priority_idx)
        area_col = get_column_letter(area_idx)
        dept_col = get_column_letter(dept_idx)
        priority_range = f"{priority_col}{data_row_start}:{priority_col}{data_row_end}"
        area_range = f"{area_col}{data_row_start}:{area_col}{data_row_end}"
        dept_range = f"{dept_col}{data_row_start}:{dept_col}{data_row_end}"

        last_data_row = len(df) + 1
        date_number_format = "yyyy-mm-dd hh:mm:ss"
        duration_number_format = "[h]:mm:ss"

        for col_name in datetime_columns:
            col_idx = df.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(col_idx)
            for row_idx in range(data_row_start, last_data_row + 1):
                ws[f"{col_letter}{row_idx}"].number_format = date_number_format

        for col_name in duration_columns:
            col_idx = df.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(col_idx)
            for row_idx in range(data_row_start, last_data_row + 1):
                ws[f"{col_letter}{row_idx}"].number_format = duration_number_format

        priority_formula = '"' + ",".join(PRIORITY_OPTIONS) + '"'
        dv_priority = DataValidation(type="list", formula1=priority_formula, allow_blank=True)
        ws.add_data_validation(dv_priority)
        dv_priority.add(priority_range)

        area_formula = '"' + ",".join(AREA_OPTIONS) + '"'
        dv_area = DataValidation(type="list", formula1=area_formula, allow_blank=True)
        ws.add_data_validation(dv_area)
        dv_area.add(area_range)

        if CBSA_DEPARTMENTS:
            validation_sheet_name = "Listas"
            if validation_sheet_name in wb.sheetnames:
                val_sheet = wb[validation_sheet_name]
                val_sheet.delete_rows(1, val_sheet.max_row)
            else:
                val_sheet = wb.create_sheet(validation_sheet_name)

            for idx, dept_value in enumerate(CBSA_DEPARTMENTS, start=1):
                val_sheet.cell(row=idx, column=1, value=dept_value)
            val_sheet.sheet_state = "hidden"

            if "CBSA_Departments" in wb.defined_names:
                del wb.defined_names["CBSA_Departments"]

            dept_range_ref = f"'{validation_sheet_name}'!$A$1:$A${len(CBSA_DEPARTMENTS)}"
            defined_name = DefinedName(name="CBSA_Departments", attr_text=dept_range_ref)
            if hasattr(wb.defined_names, "append"):
                wb.defined_names.append(defined_name)
            else:
                wb.defined_names.add(defined_name)

            dept_formula = (
                f'=IF(INDIRECT("${area_col}"&ROW())="{AREA_FONDOS}","-",'
                f'IF(INDIRECT("${area_col}"&ROW())="{AREA_CBSA}",CBSA_Departments,""))'
            )
            dv_dept = DataValidation(type="list", formula1=dept_formula, allow_blank=True)
            ws.add_data_validation(dv_dept)
            dv_dept.add(dept_range)
        else:
            log("Aviso: la lista de departamentos CBSA está vacía; no se creó la validación de datos para 'Departamento'.")

    log(f"Listo. Archivo generado: {out_path}")

if __name__ == "__main__":
    main()
