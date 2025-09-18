from datetime import datetime, time as dt_time
from pathlib import Path
import sys

import openpyxl
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import tickets_parser


def test_generated_excel_has_expected_columns(tmp_path, monkeypatch):
    output_path = tmp_path / "Tickets.xlsx"
    dummy_pdf = tmp_path / "Ticket-0001.pdf"
    dummy_pdf.write_bytes(b"")

    monkeypatch.setattr(tickets_parser, "glob", lambda pattern: [str(dummy_pdf)])

    def fake_parse_pdf(path, tz, now):
        return {
            "N° Ticket": "123",
            "Título del ticket": "Prueba",
            "Estado BW": "Abierto",
            "Prioridad": "Alta",
            "Departamento": "IT",
            "Fecha de creación": "",
            "Autor": "Alice",
            "Última respuesta por": "Bob",
            "Última respuesta el": "",
        }

    monkeypatch.setattr(tickets_parser, "parse_pdf", fake_parse_pdf)

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "tickets_parser.py",
            "--pdf_dir",
            str(tmp_path),
            "--out",
            str(output_path),
        ],
    )

    tickets_parser.main()

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook["Tickets"]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))

    expected_order = [
        "N Ticket",
        "Título del ticket",
        "Autor",
        "Prioridad",
        "Área",
        "Departamento",
    ]

    assert header_row[: len(expected_order)] == tuple(expected_order)

    priority_column_index = header_row.index("Prioridad") + 1
    priority_cell_value = worksheet.cell(row=2, column=priority_column_index).value
    assert priority_cell_value in (None, "")

    priority_column_letter = get_column_letter(priority_column_index)
    expected_priority_range = f"{priority_column_letter}2:{priority_column_letter}200"
    expected_formula = '"' + ",".join(tickets_parser.PRIORITY_OPTIONS) + '"'

    priority_validation = None
    for validation in worksheet.data_validations.dataValidation:
        if (
            str(validation.sqref) == expected_priority_range
            and validation.type == "list"
        ):
            priority_validation = validation
            break

    assert priority_validation is not None
    assert priority_validation.formula1 == expected_formula
    assert priority_validation.allow_blank


def test_reprocess_preserves_manual_columns_and_appends_new(tmp_path, monkeypatch):
    output_path = tmp_path / "Tickets.xlsx"
    pdf_one = tmp_path / "Ticket-0001.pdf"
    pdf_two = tmp_path / "Ticket-0002.pdf"
    pdf_one.write_bytes(b"")
    pdf_two.write_bytes(b"")

    def parse_first(path, tz, now):
        assert path == str(pdf_one)
        return {
            "N° Ticket": "123",
            "Título del ticket": "Inicial",
            "Estado BW": "Abierto",
            "Prioridad": "Alta",
            "Departamento": "IT",
            "Fecha de creación": "",
            "Autor": "Alice",
            "Última respuesta por": "Bob",
            "Última respuesta el": "",
        }

    monkeypatch.setattr(tickets_parser, "glob", lambda pattern: [str(pdf_one)])
    monkeypatch.setattr(tickets_parser, "parse_pdf", parse_first)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "tickets_parser.py",
            "--pdf_dir",
            str(tmp_path),
            "--out",
            str(output_path),
        ],
    )

    tickets_parser.main()

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook["Tickets"]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    ticket_idx = headers.index("N Ticket") + 1
    priority_idx = headers.index("Prioridad") + 1
    area_idx = headers.index("Área") + 1
    dept_idx = headers.index("Departamento") + 1

    worksheet.cell(row=2, column=priority_idx, value="Manual Prioridad")
    worksheet.cell(row=2, column=area_idx, value="Área Manual")
    worksheet.cell(row=2, column=dept_idx, value="Departamento Manual")

    workbook.save(output_path)
    workbook.close()

    def parse_second(path, tz, now):
        assert path in {str(pdf_one), str(pdf_two)}
        if path == str(pdf_one):
            return {
                "N° Ticket": "123",
                "Título del ticket": "Actualizado",
                "Estado BW": "Cerrado",
                "Prioridad": "Alta",
                "Departamento": "IT",
                "Fecha de creación": "",
                "Autor": "Alice",
                "Última respuesta por": "Bob",
                "Última respuesta el": "",
            }
        return {
            "N° Ticket": "456",
            "Título del ticket": "Nuevo",
            "Estado BW": "Abierto",
            "Prioridad": "Media",
            "Departamento": "Operaciones",
            "Fecha de creación": "",
            "Autor": "Carol",
            "Última respuesta por": "Dave",
            "Última respuesta el": "",
        }

    monkeypatch.setattr(
        tickets_parser,
        "glob",
        lambda pattern: [str(pdf_one), str(pdf_two)],
    )
    monkeypatch.setattr(tickets_parser, "parse_pdf", parse_second)

    tickets_parser.main()

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook["Tickets"]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    ticket_idx = headers.index("N Ticket") + 1
    title_idx = headers.index("Título del ticket") + 1

    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    row_by_ticket = {}
    for row in rows:
        ticket = row[ticket_idx - 1]
        if ticket:
            row_by_ticket.setdefault(ticket, []).append(row)

    assert "123" in row_by_ticket
    assert len(row_by_ticket["123"]) == 1

    ticket_123 = row_by_ticket["123"][0]
    assert ticket_123[title_idx - 1] == "Actualizado"
    assert ticket_123[priority_idx - 1] == "Manual Prioridad"
    assert ticket_123[area_idx - 1] == "Área Manual"
    assert ticket_123[dept_idx - 1] == "Departamento Manual"

    assert "456" in row_by_ticket
    assert len(row_by_ticket["456"]) == 1

    ticket_456 = row_by_ticket["456"][0]
    assert ticket_456[priority_idx - 1] in (None, "")
    assert ticket_456[area_idx - 1] in (None, "")
    assert ticket_456[dept_idx - 1] in (None, "")

    workbook.close()


def test_dates_are_native_and_durations_are_text(tmp_path, monkeypatch):
    output_path = tmp_path / "Tickets.xlsx"
    pdf_path = tmp_path / "Ticket-0001.pdf"
    pdf_path.write_bytes(b"")

    class FixedDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2024, 1, 10, 12, 0, tzinfo=tz)

    def fake_parse_pdf(path, tz, now):
        assert path == str(pdf_path)
        return {
            "N° Ticket": "789",
            "Título del ticket": "Fechas",
            "Estado BW": "Abierto",
            "Prioridad": "Media",
            "Departamento": "IT",
            "Fecha de creación": "09/01/2024 08:00",
            "Autor": "Eve",
            "Última respuesta por": "Frank",
            "Última respuesta el": "10/01/2024 09:30",
        }

    monkeypatch.setattr(tickets_parser, "glob", lambda pattern: [str(pdf_path)])
    monkeypatch.setattr(tickets_parser, "parse_pdf", fake_parse_pdf)
    monkeypatch.setattr(tickets_parser, "datetime", FixedDateTime)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "tickets_parser.py",
            "--pdf_dir",
            str(tmp_path),
            "--out",
            str(output_path),
        ],
    )

    tickets_parser.main()

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook["Tickets"]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    creation_idx = headers.index("Fecha de creación") + 1
    last_response_idx = headers.index("Última respuesta el") + 1
    wait_idx = headers.index("Tiempo parado desde la última respuesta") + 1
    open_idx = headers.index("Tiempo abierto (si sigue abierto)") + 1

    creation_cell = worksheet.cell(row=2, column=creation_idx)
    last_response_cell = worksheet.cell(row=2, column=last_response_idx)
    wait_cell = worksheet.cell(row=2, column=wait_idx)
    open_cell = worksheet.cell(row=2, column=open_idx)

    assert isinstance(creation_cell.value, datetime)
    assert isinstance(last_response_cell.value, datetime)
    assert creation_cell.number_format == "yyyy-mm-dd hh:mm:ss"
    assert last_response_cell.number_format == "yyyy-mm-dd hh:mm:ss"

    assert isinstance(wait_cell.value, str)
    assert isinstance(open_cell.value, str)
    assert wait_cell.number_format == "@"
    assert open_cell.number_format == "@"

    assert wait_cell.value == "0.02:30:00"
    assert open_cell.value == "1.04:00:00"

    workbook.close()


def test_existing_duration_values_are_preserved_and_normalized(tmp_path, monkeypatch):
    output_path = tmp_path / "Tickets.xlsx"
    pdf_path = tmp_path / "Ticket-0003.pdf"
    pdf_path.write_bytes(b"")

    headers = [
        "N Ticket",
        "Título del ticket",
        "Autor",
        "Prioridad",
        "Área",
        "Departamento",
        "Fecha de creación",
        "Última respuesta por",
        "Última respuesta el",
        "Tiempo parado desde la última respuesta",
        "Tiempo abierto (si sigue abierto)",
    ]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Tickets"
    worksheet.append(headers)
    worksheet.append(
        [
            "123",
            "Antiguo 1",
            "Alice",
            "",
            "",
            "",
            "",
            "",
            "",
            "18d 16h",
            "5d 3h",
        ]
    )
    worksheet.append(
        [
            "456",
            "Antiguo 2",
            "Bob",
            "",
            "",
            "",
            "",
            "",
            "",
            1.5,
            dt_time(10, 0),
        ]
    )
    workbook.save(output_path)
    workbook.close()

    monkeypatch.setattr(tickets_parser, "glob", lambda pattern: [str(pdf_path)])

    def fake_parse_pdf(path, tz, now):
        assert path == str(pdf_path)
        return {
            "N° Ticket": "789",
            "Título del ticket": "Nuevo",
            "Estado BW": "Abierto",
            "Prioridad": "Alta",
            "Departamento": "IT",
            "Fecha de creación": "08/20/2025 08:00 AM",
            "Autor": "Carol",
            "Última respuesta por": "Dave",
            "Última respuesta el": "",
        }

    monkeypatch.setattr(tickets_parser, "parse_pdf", fake_parse_pdf)

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "tickets_parser.py",
            "--pdf_dir",
            str(tmp_path),
            "--out",
            str(output_path),
        ],
    )

    tickets_parser.main()

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook["Tickets"]

    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    wait_idx = headers.index("Tiempo parado desde la última respuesta") + 1
    open_idx = headers.index("Tiempo abierto (si sigue abierto)") + 1

    results = {}
    for row in worksheet.iter_rows(min_row=2):
        ticket = row[0].value
        wait_cell = row[wait_idx - 1]
        open_cell = row[open_idx - 1]
        results[ticket] = (
            wait_cell.value,
            wait_cell.number_format,
            open_cell.value,
            open_cell.number_format,
        )

    workbook.close()

    assert results["123"] == ("18.16:00:00", "@", "5.03:00:00", "@")
    assert results["456"] == ("1.12:00:00", "@", "0.10:00:00", "@")
    assert "789" in results
