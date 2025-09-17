import subprocess
import sys
from pathlib import Path

import pytest

pytest.importorskip("pandas")
pytest.importorskip("openpyxl")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _count_data_rows(ws):
    max_col = ws.max_column
    data_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
        if all(value in (None, "") for value in row_values):
            break
        data_rows += 1
    return data_rows


def test_department_cells_blank_and_validated(tmp_path):
    repo_root = Path(__file__).resolve().parents[1]
    pdf_dir = repo_root / "reports"
    out_path = tmp_path / "Tickets.xlsx"

    result = subprocess.run(
        [
            sys.executable,
            str(repo_root / "tickets_parser.py"),
            "--pdf_dir",
            str(pdf_dir),
            "--pattern",
            "Ticket-115249.pdf",
            "--out",
            str(out_path),
        ],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )

    assert out_path.exists(), result.stderr

    wb = load_workbook(out_path)
    ws = wb["Tickets"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Departamento" in headers
    dept_col_idx = headers.index("Departamento") + 1
    dept_letter = get_column_letter(dept_col_idx)

    data_rows = _count_data_rows(ws)
    assert data_rows > 0

    for row_idx in range(2, 2 + data_rows):
        cell_value = ws.cell(row=row_idx, column=dept_col_idx).value
        assert cell_value in (None, "")

    expected_end = max(data_rows + 1, 200)
    expected_sqref = f"{dept_letter}2:{dept_letter}{expected_end}"

    dept_validations = [
        dv
        for dv in ws.data_validations.dataValidation
        if str(dv.sqref) == expected_sqref
    ]

    assert dept_validations, f"No data validation for Departamento range (found: {[str(dv.sqref) for dv in ws.data_validations.dataValidation]})"

    dv = dept_validations[0]
    assert dv.type == "list"
    assert dv.allow_blank
