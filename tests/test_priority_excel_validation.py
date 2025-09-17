from datetime import datetime
from zoneinfo import ZoneInfo

import pandas as pd
import pytest

pytest.importorskip("openpyxl")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import tickets_parser as tp


def test_parse_pdf_leaves_priority_blank(monkeypatch):
    tz = ZoneInfo("UTC")
    now = datetime(2024, 1, 1, 12, 0, tzinfo=tz)

    sample_text = "X" * 120

    monkeypatch.setattr(tp, "extract_text", lambda path: sample_text)
    monkeypatch.setattr(tp, "clean_text", lambda text: text)

    monkeypatch.setattr(
        tp,
        "parse_header_block",
        lambda cleaned, lines: {
            "Ticket Number": "123",
            "Status": "Abierto",
            "Priority": "Alta",
            "Department": "IT",
            "Create Date": "01/01/2024 10:00 AM",
        },
    )

    monkeypatch.setattr(tp, "extract_title_after_urgency", lambda cleaned: "Título")
    entries = [
        {"author": "Alice", "stamp": "01/01/2024 10:00 AM", "body": "Mensaje"},
        {"author": "Bob", "stamp": "01/01/2024 11:00 AM", "body": "Respuesta"},
    ]
    monkeypatch.setattr(tp, "extract_thread_entries", lambda cleaned: entries)
    monkeypatch.setattr(tp, "is_auto", lambda author, body: False)
    monkeypatch.setattr(tp, "fallback_last_author_from_tail", lambda cleaned: "Fallback")

    row = tp.parse_pdf("dummy.pdf", tz, now)

    assert row["Prioridad"] == ""
    assert row["Departamento"] == "IT"


def test_priority_column_validation_and_blank(tmp_path):
    columns = [
        "N° Ticket",
        "Título del ticket",
        "Estado",
        "Prioridad",
        "Departamento",
        "Fecha de creación",
        "Autor",
        "Última respuesta por",
        "Última respuesta el",
        "Error",
        "Área",
    ]

    df = pd.DataFrame([
        {
            "N° Ticket": "123",
            "Título del ticket": "Título",
            "Estado": "Abierto",
            "Prioridad": "",
            "Departamento": "IT",
            "Fecha de creación": "",
            "Autor": "Alice",
            "Última respuesta por": "Bob",
            "Última respuesta el": "",
            "Error": "",
            "Área": tp.AREA_CBSA,
        }
    ], columns=columns)

    out_path = tmp_path / "out.xlsx"
    data_row_start = 2
    data_row_end = max(len(df) + 1, 200)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tickets")
        wb = writer.book
        ws = writer.sheets["Tickets"]
        tp._apply_excel_validations(df, wb, ws, data_row_start, data_row_end)

    wb = load_workbook(out_path)
    ws = wb["Tickets"]

    priority_col_idx = df.columns.get_loc("Prioridad") + 1
    priority_col_letter = get_column_letter(priority_col_idx)
    priority_cell = ws.cell(row=data_row_start, column=priority_col_idx)

    assert priority_cell.value in (None, "")

    expected_range = f"{priority_col_letter}{data_row_start}:{priority_col_letter}{data_row_end}"
    formula = '"' + ",".join(tp.PRIORITY_OPTIONS) + '"'

    priority_dv = None
    for dv in ws.data_validations.dataValidation:
        if expected_range in str(dv.sqref):
            priority_dv = dv
            break

    assert priority_dv is not None
    assert priority_dv.formula1 == formula
    assert priority_dv.allow_blank
